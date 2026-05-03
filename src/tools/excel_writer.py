"""ExcelWriterTool — Escribe datos estructurados a .xlsx local.

Complemento de ExcelReaderTool. Permite a agentes generar presupuestos,
órdenes de compra, y otros documentos como archivos Excel.
Reemplazable por Google Sheets API sin cambiar interfaz del agente.
"""

from __future__ import annotations

import json
from pathlib import Path
from typing import Type

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill
from pydantic import BaseModel, Field

from src.tools.base_tool import OrgBaseTool
from src.tools.registry import register_tool

BASE_DIR = Path(__file__).resolve().parent.parent.parent / "PROJECT-Aybar"


class ExcelWriterInput(BaseModel):
    filename: str = Field(description="Nombre del archivo .xlsx a crear/actualizar")
    sheet_name: str = Field(default="Datos", description="Nombre de la sheet")
    data: str = Field(description="JSON array de objetos a escribir como filas")
    mode: str = Field(default="overwrite", description="overwrite | append")


@register_tool(
    "excel_writer",
    description="Escribe datos estructurados a archivos Excel del proyecto Aybar.",
    tags=["business", "excel", "aybar"],
)
class ExcelWriterTool(OrgBaseTool):
    name: str = "excel_writer"
    description: str = (
        "Escribe un JSON array como sheet de Excel en PROJECT-Aybar/. "
        "Los objetos del array se convierten en filas con cabeceras como columnas. "
        "Soporta modo overwrite (reemplaza) o append (agrega filas)."
    )
    args_schema: Type[BaseModel] = ExcelWriterInput

    def _run(self, filename: str, sheet_name: str = "Datos", data: str = "[]", mode: str = "overwrite") -> str:
        filepath = BASE_DIR / filename

        try:
            records = json.loads(data)
        except json.JSONDecodeError as e:
            return json.dumps({"error": f"JSON inválido: {e}"})

        if not isinstance(records, list):
            return json.dumps({"error": "data debe ser un JSON array"})

        if not records:
            return json.dumps({"error": "data vacío"})

        try:
            if filepath.exists() and mode == "append":
                wb = openpyxl.load_workbook(filepath)
            else:
                wb = openpyxl.Workbook()

            if sheet_name in wb.sheetnames:
                ws = wb[sheet_name]
                if mode == "overwrite":
                    ws.delete_rows(1, ws.max_row)
                    start_row = 1
                else:
                    start_row = ws.max_row + 1 if ws.max_row else 1
            else:
                ws = wb.create_sheet(title=sheet_name) if len(wb.sheetnames) > 0 else wb.active
                ws.title = sheet_name
                start_row = 1

            headers = list(records[0].keys())
            if start_row == 1:
                for col, h in enumerate(headers, 1):
                    cell = ws.cell(row=1, column=col, value=h)
                    cell.font = Font(bold=True)
                    cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
                    cell.font = Font(bold=True, color="FFFFFF")
                    cell.alignment = Alignment(horizontal="center")

            for i, record in enumerate(records, start=start_row + (1 if start_row == 1 else 0)):
                for col, h in enumerate(headers, 1):
                    val = record.get(h, "")
                    if isinstance(val, float):
                        cell = ws.cell(row=i, column=col, value=val)
                        cell.number_format = '#,##0.00'
                    else:
                        ws.cell(row=i, column=col, value=str(val) if val is not None else "")

            wb.save(filepath)
            wb.close()

            return json.dumps({
                "status": "ok",
                "filename": filename,
                "sheet": sheet_name,
                "rows_written": len(records),
                "mode": mode,
                "path": str(filepath),
            }, ensure_ascii=False)

        except Exception as e:
            return json.dumps({"error": f"Error escribiendo Excel: {e}"})
