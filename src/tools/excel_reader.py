"""ExcelReaderTool — Lee archivos .xlsx locales y retorna JSON estructurado.

Útil para agentes que necesitan consultar datos de negocio en sheets
sin integración externa. Reemplazable por Google Sheets API sin
cambiar la interfaz del agente.
"""

from __future__ import annotations

import json
from pathlib import Path
from typing import Any, Dict, List, Optional, Type

import openpyxl
from pydantic import BaseModel, Field

from src.tools.base_tool import OrgBaseTool
from src.tools.registry import register_tool

BASE_DIR = Path(__file__).resolve().parent.parent.parent / "PROJECT-Aybar"


class ExcelReaderInput(BaseModel):
    filename: str = Field(description="Nombre del archivo .xlsx (ej: precios_bebidas.xlsx)")
    sheet_name: Optional[str] = Field(None, description="Nombre de la sheet (opcional). Si no se especifica, retorna todas las sheets.")


@register_tool(
    "excel_reader",
    description="Lee archivos Excel del proyecto Aybar y retorna datos estructurados en JSON.",
    tags=["business", "excel", "aybar"],
)
class ExcelReaderTool(OrgBaseTool):
    name: str = "excel_reader"
    description: str = (
        "Lee archivos .xlsx de la carpeta PROJECT-Aybar y retorna "
        "los datos como JSON estructurado. Cada sheet se convierte en "
        "una lista de diccionarios con cabeceras como keys."
    )
    args_schema: Type[BaseModel] = ExcelReaderInput

    def _run(self, filename: str, sheet_name: Optional[str] = None) -> str:
        filepath = BASE_DIR / filename
        if not filepath.exists():
            return json.dumps({"error": f"Archivo '{filename}' no encontrado en {BASE_DIR}"})
        if filepath.suffix.lower() not in (".xlsx", ".xls"):
            return json.dumps({"error": f"Formato no soportado: {filepath.suffix}"})

        try:
            wb = openpyxl.load_workbook(filepath, read_only=True, data_only=True)
        except Exception as e:
            return json.dumps({"error": f"Error abriendo archivo: {e}"})

        result: Dict[str, List[Dict[str, Any]]] = {}

        sheets = [sheet_name] if sheet_name else wb.sheetnames

        for sn in sheets:
            if sn not in wb.sheetnames:
                result[sn] = [{"error": f"Sheet '{sn}' no encontrada"}]
                continue

            ws = wb[sn]
            rows = list(ws.iter_rows(values_only=True))
            if not rows:
                result[sn] = []
                continue

            # Find header row (first row with meaningful content)
            header_row_idx = 0
            for i, row in enumerate(rows):
                vals = [v for v in row if v is not None]
                if len(vals) >= 2:
                    header_row_idx = i
                    break

            headers = [str(v) if v is not None else None for v in rows[header_row_idx]]
            data = []
            for row in rows[header_row_idx + 1:]:
                vals = [v for v in row]
                if all(v is None for v in vals):
                    continue
                row_dict = {}
                for idx, h in enumerate(headers):
                    if h and idx < len(vals):
                        val = vals[idx]
                        if isinstance(val, (int, float)):
                            row_dict[h] = val
                        elif val is not None:
                            row_dict[h] = str(val)
                if row_dict:
                    data.append(row_dict)

            result[sn] = data

        wb.close()
        return json.dumps(result, ensure_ascii=False, default=str)
